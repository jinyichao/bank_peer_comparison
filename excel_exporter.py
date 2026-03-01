import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")   # dark blue
_ALT_FILL    = PatternFill(fill_type="solid", fgColor="D6E4F0")   # light blue
_WHITE_FILL  = PatternFill(fill_type="solid", fgColor="FFFFFF")
_THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def to_excel_bytes(df: pd.DataFrame, bullets: list[str] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Comparison"

    # ── Header row ──────────────────────────────────────────────────────────
    header = ["Metric"] + list(df.columns)
    for col_idx, value in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, (label, row_data) in enumerate(df.iterrows(), start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL

        # Metric label (first column)
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(bold=True, size=10)
        label_cell.fill = fill
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.border = _THIN_BORDER

        # Bank value columns
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = _THIN_BORDER
            cell.font = Font(size=10)

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 42
    for col_idx in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # ── Row height ───────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(df) + 2):
        ws.row_dimensions[row_idx].height = 20

    # ── Freeze top row + left column ─────────────────────────────────────────
    ws.freeze_panes = "B2"

    # ── Summary section ───────────────────────────────────────────────────────
    if bullets:
        num_cols = len(header)
        summary_start_row = len(df) + 3   # blank row gap after data

        # "Summary" header
        summary_header = ws.cell(row=summary_start_row, column=1, value="Summary")
        summary_header.font = Font(bold=True, color="FFFFFF", size=11)
        summary_header.fill = _HEADER_FILL
        summary_header.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(
            start_row=summary_start_row, start_column=1,
            end_row=summary_start_row, end_column=num_cols,
        )
        ws.row_dimensions[summary_start_row].height = 24

        # Bullet rows
        for j, bullet in enumerate(bullets):
            r = summary_start_row + 1 + j
            cell = ws.cell(row=r, column=1, value=f"• {bullet}")
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            fill = _ALT_FILL if j % 2 == 0 else _WHITE_FILL
            cell.fill = fill
            ws.merge_cells(
                start_row=r, start_column=1,
                end_row=r, end_column=num_cols,
            )
            ws.row_dimensions[r].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
